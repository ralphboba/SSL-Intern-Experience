#안녕하세요! 해당 코드를 작성한 SSL 두은혁 인턴입니다. 해당 코드 사용방법을 알려드리겠습니다. 원하시는 테이블의 CSS selector 코드를 찾으신 다음에
#15번째 줄 괄호 안에 있는 정보를 해당 CSS selector 코드로 바꿔주시면 됩니다! 그리고 td와 th의 정보가 같이 추출 되는데 th 또는 td 정보만 원하신다면 26번째 줄에서 수정하시면 됩니다.


import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os

def get_table_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # 테이블 선택
    table = soup.select_one('#view02')  #바꾸시면 됩니다
    if not table:
        print(f"테이블을 찾을 수 없습니다: {url}")
        return []  # 데이터가 없는 경우 빈 리스트 반환
    
    rows = table.find_all('tr')
    
    # 테이블 데이터 저장 리스트
    table_data = []
    for row in rows:
        row_data = []
        cols = row.find_all(['td' , 'th']) 
        for col in cols:
            cell_data = col.get_text(strip=True)
            colspan = int(col.get('colspan', 1))
            rowspan = int(col.get('rowspan', 1))
            row_data.append((cell_data, rowspan, colspan))
        table_data.append(row_data)
    
    return table_data

def write_to_excel(table_data, start_row, ws, url):
    if not table_data:
        ws.cell(row=start_row, column=1, value="정보가 없습니다")
        ws.cell(row=start_row, column=2, value=url)
        return start_row + 1
    
    max_cols = max(sum(cell[2] for cell in row) for row in table_data)
    max_rows = len(table_data)
    cell_matrix = [["" for _ in range(max_cols)] for _ in range(max_rows)]
    
    for r_idx, row in enumerate(table_data):
        c_idx = 0
        for cell in row:
            cell_value, rowspan, colspan = cell
            while c_idx < max_cols and cell_matrix[r_idx][c_idx] != "":
                c_idx += 1
            for r in range(rowspan):
                for c in range(colspan):
                    if r_idx + r < max_rows and c_idx + c < max_cols:
                        cell_matrix[r_idx + r][c_idx + c] = cell_value
            c_idx += colspan

    current_row = start_row
    for r_idx, row in enumerate(cell_matrix):
        if all(cell == "" for cell in row):  # 빈 행은 건너뛰기
            continue
        
        for c_idx, cell_value in enumerate(row, 3):  # C열부터 시작
            ws.cell(row=current_row, column=c_idx, value=cell_value)
            # URL을 바로 옆 셀에 추가
            ws.cell(row=current_row, column=c_idx + 1, value=url)
        
        current_row += 1  # 빈 행이 아닌 경우에만 다음 행으로 이동
    
    return current_row  # 다음 데이터 시작 행 반환

# URL 리스트
urls = [
"https://oasis.kiom.re.kr/oasis/herb/monoDetailView_M05.jsp?idx=1&tab=5#view02",
"https://oasis.kiom.re.kr/oasis/herb/monoDetailView_M05.jsp?idx=2&tab=5#view02",
"https://oasis.kiom.re.kr/oasis/herb/monoDetailView_M05.jsp?idx=3&tab=5#view02"
]

# 워크북 생성
wb = Workbook()
ws = wb.active

# 각 URL에서 데이터 추출 및 엑셀 파일로 저장
start_row = 1
for i, url in enumerate(urls):
    table_data = get_table_data(url)
    start_row = write_to_excel(table_data, start_row, ws, url)

# 엑셀 파일 저장
output_file = "output1.xlsx"
wb.save(output_file)

# 엑셀 파일 자동 열기
os.system(f'start excel.exe "{output_file}"')
